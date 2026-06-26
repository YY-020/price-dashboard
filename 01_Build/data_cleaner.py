"""
数据预处理脚本 - 从 00_Inbox 读取原始 Excel，清洗后输出 CSV 到 02_Output
"""
import os
import re
import glob
import pandas as pd
import openpyxl
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INBOX_DIR = PROJECT_ROOT / "00_Inbox"
OUTPUT_DIR = PROJECT_ROOT / "02_Output"
OUTPUT_CSV = OUTPUT_DIR / "price_data_clean.csv"

# 材料分组定义（与 data_loader.py 保持一致，用于去重匹配）
MATERIAL_GROUPS = {
    "上游原材料": ["乙烯", "乙炔", "硫酸镍", "硫酸钴", "硫酸锰", "碳酸锂", "氢氧化锂·微粉", "氢氧化锂·粗颗粒"],
    "电解液溶剂": ["碳酸乙烯酯EC", "碳酸丙烯酯PC", "碳酸二甲酯DMC", "碳酸甲乙酯EMC", "碳酸二乙酯DEC", "乙酸甲酯MA"],
    "电解液添加剂": ["碳酸亚乙烯酯（百川）", "碳酸亚乙烯酯（有色网）", "氟代碳酸乙烯酯FEC", "硫酸乙烯酯DTD", "丙烯磺酸内酯PRS", "1,3-丙烷磺酸内酯PS"],
    "电解液锂盐": ["LiFSi·折固", "LiFSi·固体", "六氟磷酸锂LiFP6", "LiDFOB·固体", "LiPO2F2·固体"],
    "电解液": ["电解液·三元用", "电解液·磷酸铁锂用"],
    "磷酸铁": ["磷酸铁"],
    "磷酸铁锂": ["磷酸铁锂·储能高端型", "磷酸铁锂·动力高端型"],
    "碳纳米管": ["多壁浆料-高端", "多壁浆料-低端", "多壁粉体-高端", "多壁粉体-低端", "单壁粉体-高端", "单壁粉体-低端"],
    "NMP": ["NMP"],
    "电芯": ["8系三元电芯", "方形铁锂储能型电芯", "方形铁锂动力型电芯"],
}


def find_latest_excel() -> Path:
    """按文件名日期识别最新的 Excel 文件"""
    pattern = str(INBOX_DIR / "price_data_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"在 {INBOX_DIR} 中未找到 price_data_*.xlsx 文件")

    # 尝试从文件名提取日期排序
    def extract_date(f):
        match = re.search(r"(\d{4})", os.path.basename(f))
        return match.group(1) if match else "0000"

    files.sort(key=extract_date, reverse=True)
    return Path(files[0])


def _resolve_dup_name(name: str, spec: str, source: str) -> str:
    """对重复材料名，用规格+数据源匹配 MATERIAL_GROUPS 中的标准名"""
    candidates = []
    for group_name, members in MATERIAL_GROUPS.items():
        for member in members:
            if name in member and name != member:
                candidates.append(member)

    # 用规格关键词匹配
    if spec and spec != "/":
        for c in candidates:
            for keyword in spec.replace(",", " ").split():
                if keyword in c:
                    return c

    # 用数据源关键词匹配
    if source:
        for c in candidates:
            if source in c:
                return c

    # 回退：用规格后缀
    if spec and spec != "/":
        return f"{name}·{spec}"

    # 回退：用数据源后缀
    if source:
        return f"{name}（{source}）"

    return name


def clean_excel(filepath: Path) -> None:
    """读取原始 Excel，清洗后输出 CSV"""
    print(f"读取: {filepath}")
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    # 1. 处理合并单元格：先取消合并，再把分组标签填充到每一列
    merged = list(ws.merged_cells.ranges)
    # 先收集所有合并区域的值和范围，再统一取消合并
    merge_info = []
    for mr in merged:
        val = ws.cell(mr.min_row, mr.min_col).value
        merge_info.append((mr.min_row, mr.min_col, mr.max_col, val))
    # 取消所有合并
    for mr in merged:
        ws.unmerge_cells(str(mr))
    # 填充值
    for row, min_col, max_col, val in merge_info:
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).value = val

    # 2. 找到第2行中非空的列（有效列）
    valid_cols = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(2, col).value
        if val is not None:
            valid_cols.append(col)

    print(f"有效列数: {len(valid_cols)} (列索引: {valid_cols[0]}-{valid_cols[-1]})")

    # 3. 读取表头（前7行）
    header_rows = {}
    for row_idx in range(1, 8):
        row_data = {}
        for col in valid_cols:
            row_data[col] = ws.cell(row_idx, col).value
        header_rows[row_idx] = row_data

    # 4. 读取数据区（第8行开始，自动检测结束）
    data_start = 8
    records = []
    for row_idx in range(data_start, ws.max_row + 1):
        date_val = ws.cell(row_idx, valid_cols[0]).value
        if date_val is None:
            break
        row_data = {}
        for col in valid_cols:
            row_data[col] = ws.cell(row_idx, col).value
        records.append(row_data)

    print(f"数据行数: {len(records)}")

    # 5. 构建 DataFrame
    # 列名：用第2行材料名，对重复名用规格/数据源区分
    from collections import Counter
    raw_names = [str(header_rows[2].get(col, "")).strip() for col in valid_cols[1:]]
    name_counts = Counter(raw_names)

    col_names = {}
    for col in valid_cols[1:]:
        name = str(header_rows[2].get(col, "")).strip()
        if name_counts[name] > 1:
            spec = str(header_rows[3].get(col, "") or "").strip()
            source = str(header_rows[6].get(col, "") or "").strip()
            name = _resolve_dup_name(name, spec, source)
        col_names[col] = name

    df_data = {}
    # 第1列是日期
    date_col = valid_cols[0]
    dates = [r[date_col] for r in records]
    df_data["日期"] = dates

    for col in valid_cols[1:]:
        col_name = col_names[col]
        values = [r[col] for r in records]
        df_data[col_name] = values

    df = pd.DataFrame(df_data)
    df["日期"] = pd.to_datetime(df["日期"])
    df = df.sort_values("日期").reset_index(drop=True)

    # 6. 价格列强制转数值
    price_cols = [c for c in df.columns if c != "日期"]
    for col in price_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # 7. 保存元信息（前6行表头）为单独的 CSV
    meta_rows = []
    row_labels = ["分组", "材料", "规格", "CAS", "单位", "数据源", "化学式"]
    for i, label in enumerate(row_labels):
        row_data = {"行标签": label}
        for col in valid_cols[1:]:
            col_name = col_names[col]
            row_data[col_name] = header_rows[i + 1].get(col, "")
        meta_rows.append(row_data)
    df_meta = pd.DataFrame(meta_rows)
    meta_path = OUTPUT_DIR / "price_data_meta.csv"
    df_meta.to_csv(str(meta_path), index=False, encoding="utf-8-sig")
    print(f"元信息已保存: {meta_path}")

    # 8. 保存清洗后的数据
    df.to_csv(str(OUTPUT_CSV), index=False, encoding="utf-8-sig")
    print(f"清洗数据已保存: {OUTPUT_CSV}")
    print(f"最终 DataFrame: {df.shape[0]} 行 x {df.shape[1]} 列")
    print(f"日期范围: {df['日期'].min().strftime('%Y-%m-%d')} ~ {df['日期'].max().strftime('%Y-%m-%d')}")


if __name__ == "__main__":
    excel_path = find_latest_excel()
    clean_excel(excel_path)
